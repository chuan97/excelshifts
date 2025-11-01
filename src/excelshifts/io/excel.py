"""Module to handle input and output of excel files"""

import shutil

import pandas as pd
from openpyxl import load_workbook

import excelshifts.state as state


def load_residents(
    file_path: str, sheet_name: str, start: int, n_residents: int
) -> tuple[state.Resident, ...]:
    """loads resident names & ranks from columns A & B.

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to load data from
        start: The starting row index for the residents
        n_residents: The number of residents to load

    returns:
        A tuple of Resident objects
    """

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine="openpyxl")

    ROW_BOUNDS = (start - 1, start + n_residents - 2)

    residents = []
    for row_pos in range(df.shape[0]):
        row = df.iloc[row_pos]
        if is_rowcol_in_bounds(row_pos, ROW_BOUNDS):
            if pd.notna(row[0]):
                lastrank = row[0]
            residents.append(state.Resident(row[1], lastrank))
    return tuple(residents)


def load_days(
    file_path: str, sheet_name: str, start: int, n_days: int
) -> tuple[state.Day, ...]:
    """loads day numbers and weekdays from rows 2 & 3.

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to load data from
        start: The starting column index for the days
        n_days: The number of days to load

    returns:
        A tuple of Day objects
    """

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine="openpyxl")

    COL_BOUNDS = (start - 1, start + n_days - 1)

    day_numbers = df.iloc[1, COL_BOUNDS[0] : COL_BOUNDS[1]].dropna().tolist()
    weekdays = df.iloc[2, COL_BOUNDS[0] : COL_BOUNDS[1]].dropna().tolist()

    days = [
        state.Day(number, weekday.strip())
        for number, weekday in zip(day_numbers, weekdays)
    ]

    return tuple(days)


def load_restrictions(
    file_path: str,
    sheet_name: str,
    types: list[str],
    row_start: int,
    col_start: int,
    n_residents: int,
    n_days: int,
) -> tuple[tuple[int, int], ...]:
    """loads restrictions from the Excel file.

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to load data from
        types: The types of restrictions
        row_start: The starting row index for the restrictions
        col_start: The starting column index for the restrictions
        n_residents: The number of residents
        n_days: The number of days

    returns:
        A tuple of restricted (resident_index, day_index) tuples with the restrictions
    """

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine="openpyxl")

    ROW_OFFSET = row_start - 1
    COL_OFFSET = col_start - 1

    ROW_BOUNDS = (ROW_OFFSET, n_residents + ROW_OFFSET - 1)
    COL_BOUNDS = (COL_OFFSET, n_days + COL_OFFSET - 1)

    positions = [
        (row_pos - ROW_OFFSET, col_pos - COL_OFFSET)
        for row_pos in range(df.shape[0])
        for col_pos, cell in enumerate(df.iloc[row_pos])
        if is_cell_in_bounds(row_pos, col_pos, ROW_BOUNDS, COL_BOUNDS) and cell in types
    ]
    return tuple(positions)


def load_external_rotations(
    file_path: str,
    sheet_name: str,
    row_start: int,
    col_start: int,
    n_residents: int,
    n_days: int,
) -> frozenset[int]:
    """loads external rotations from the Excel file.

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to load data from
        row_start: The starting row index for the restrictions
        col_start: The starting column index for the restrictions
        n_residents: The number of residents
        n_days: The number of days

    returns:
        A frozenset of residents who are in external rotations
    """

    e_positions = load_restrictions(
        file_path, sheet_name, ["E"], row_start, col_start, n_residents, n_days
    )

    return frozenset(i for i, _ in e_positions)


def load_preset_shifts(
    file_path: str,
    sheet_name: str,
    row_start: int,
    col_start: int,
    n_residents: int,
    n_days: int,
) -> tuple[tuple[int, int, int], ...]:
    """loads preset shifts from the Excel file.

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to load data from
        row_start: The starting row index for the preset shifts
        col_start: The starting column index for the preset shifts
        n_residents: The number of residents
        n_days: The number of days

    returns:
        A tuple of (resident_index, day_index, shift_index) tuples with the preset shifts
    """

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine="openpyxl")

    ROW_OFFSET = row_start - 1
    COL_OFFSET = col_start - 1

    ROW_BOUNDS = (ROW_OFFSET, n_residents + ROW_OFFSET - 1)
    COL_BOUNDS = (COL_OFFSET, n_days + COL_OFFSET - 1)

    positions = [
        (row_pos - ROW_OFFSET, col_pos - COL_OFFSET, state.ShiftType[cell].value)
        for row_pos in range(df.shape[0])
        for col_pos, cell in enumerate(df.iloc[row_pos])
        if is_cell_in_bounds(row_pos, col_pos, ROW_BOUNDS, COL_BOUNDS)
        and cell in list(state.ShiftType.__members__)
    ]
    return tuple(positions)


def is_cell_in_bounds(
    row_idx: int, col_idx: int, row_bounds: tuple[int, int], col_bounds: tuple[int, int]
) -> bool:
    """Check if a cell is in the bounds of the table

    args:
        row_idx: The row index of the cell
        col_idx: The column index of the cell
        row_bounds: The bounds of the rows
        col_bounds: The bounds of the columns

    returns:
        True if the cell is in the bounds, False otherwise
    """
    return is_rowcol_in_bounds(row_idx, row_bounds) and is_rowcol_in_bounds(
        col_idx, col_bounds
    )


def is_rowcol_in_bounds(idx: int, bounds: tuple[int, int]) -> bool:
    """Check if a row or column is in the bounds of the table

    args:
        idx: The row or column index
        bounds: The bounds of the row or column

    returns:
        True if the row or column is in the bounds, False otherwise
    """
    return bounds[0] <= idx <= bounds[1]


def load_totals(
    file_path: str, sheet_name: str, row_start: int, col_start: int, n_residents: int
) -> list[list[int]]:
    """loads totals from the Excel file.

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to load data from
        row_start: The starting row index for the totals
        col_start: The starting column index for the totals
        n_residents: The number of residents

    returns:
        A matrix of total shifts of each type for each resident, rows are residents, columns are shift types
    """

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine="openpyxl")

    ROW_BOUNDS = (row_start - 1, row_start + n_residents - 2)
    COL_BOUNDS = (col_start - 1, col_start + len(state.ShiftType) - 2)

    totals = []
    for row_pos in range(df.shape[0]):
        row = df.iloc[row_pos]
        if is_rowcol_in_bounds(row_pos, ROW_BOUNDS):
            shifts = []
            for col_idx, cell in enumerate(row):
                if is_rowcol_in_bounds(col_idx, COL_BOUNDS):
                    shifts.append(cell)
            totals.append(shifts)
    return totals


def copy_excel_file(original_path: str, fname_extension: str):
    """Copies an Excel file and saves it with a new filename in the same directory.

    args:
        original_path: The path to the original Excel file
        fname_extension: A string to add to the original filename

    returns:
        The path of the copied file
    """
    # Define the new file path
    new_path = original_path[:-5] + fname_extension + ".xlsx"

    # Copy the file
    shutil.copy2(original_path, new_path)

    return new_path  # Return the path of the copied file


def save_shifts(
    file_path: str,
    sheet_name: str,
    shift_matrix: list[list[str]],
    row_start: int,
    col_start: int,
):
    """Introduces the shifts into a given Excel file

    args:
        file_path: The path to the Excel file
        sheet_name: The name of the sheet to save data to
        shift_matrix: The matrix of shifts to save
        row_start: The starting row index for the shifts
        col_start: The starting column index for the shifts
    """

    # Load the existing workbook
    wb = load_workbook(file_path)

    # Select the sheet
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    sheet = wb[sheet_name]

    ROW_OFFSET = row_start - 1
    COL_OFFSET = col_start - 1

    for i, row in enumerate(shift_matrix):
        for j, shift in enumerate(row):
            if shift:
                sheet.cell(
                    row=i + ROW_OFFSET + 1, column=j + COL_OFFSET + 1, value=shift
                )

    wb.save(file_path)


# TODO: Add function to load totals of Sabados and Viernes-Domingos
# TODO: Add function to save updated totals


# Helper to load a complete Instance from Excel using the above loaders
def load_instance(
    file_path: str,
    sheet_name: str,
    *,
    residents_start: int,
    n_residents: int,
    days_start: int,
    n_days: int,
    grid_row_start: int,
    grid_col_start: int,
    p_days: list[int],
) -> state.Instance:
    """Load a complete Instance from a single Excel sheet.

    Parameters
    ----------
    residents_start: first row of the residents block (names & ranks in cols A/B)
    days_start: first column of the days block (day numbers / weekdays)
    grid_row_start, grid_col_start: top-left of the assignment grid (restrictions/presets)
    n_residents, n_days: sizes of the grid
    """
    residents = load_residents(file_path, sheet_name, residents_start, n_residents)
    days = load_days(file_path, sheet_name, days_start, n_days)

    # Restrictions & presets within the grid
    v_positions = load_restrictions(
        file_path,
        sheet_name,
        ["V"],
        grid_row_start,
        grid_col_start,
        n_residents,
        n_days,
    )
    u_positions = load_restrictions(
        file_path,
        sheet_name,
        ["U"],
        grid_row_start,
        grid_col_start,
        n_residents,
        n_days,
    )
    ut_positions = load_restrictions(
        file_path,
        sheet_name,
        ["UT"],
        grid_row_start,
        grid_col_start,
        n_residents,
        n_days,
    )
    p_positions = load_restrictions(
        file_path,
        sheet_name,
        ["P"],
        grid_row_start,
        grid_col_start,
        n_residents,
        n_days,
    )
    external_rotations = load_external_rotations(
        file_path, sheet_name, grid_row_start, grid_col_start, n_residents, n_days
    )
    presets = load_preset_shifts(
        file_path, sheet_name, grid_row_start, grid_col_start, n_residents, n_days
    )

    return state.Instance(
        residents=residents,
        days=days,
        v_positions=v_positions,
        u_positions=u_positions,
        ut_positions=ut_positions,
        p_positions=p_positions,
        extra_p_days=tuple(p_days),
        external_rotations=external_rotations,
        presets=presets,
    )
